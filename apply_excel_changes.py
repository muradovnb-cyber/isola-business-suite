"""
apply_excel_changes.py
Применяет изменения из ISOLA_Справочник_v2.xlsx в data.json:
1. Обновляет / добавляет сервисных контрагентов (лист "Услуги (подрядчики)")
2. Обновляет поле who у статей расходов на основе матрицы ролей
"""
import json, sys, shutil
sys.stdout.reconfigure(encoding='utf-8')

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl"); sys.exit(1)

EXCEL = r"C:\Users\user\Desktop\ISOLA_Справочник_v2.xlsx"
DATA  = r"C:\Users\user\клод код\таблицы экзель\ISOLA_Business_Suite\isola-repo\data.json"

wb = openpyxl.load_workbook(EXCEL, data_only=True)

with open(DATA, 'r', encoding='utf-8-sig') as f:
    d = json.load(f)

# ── 1. УСЛУГИ (подрядчики) ────────────────────────────────────────────────────
# Текущие сервисные CP (по имени → id)
svc_by_name = {c['n'].strip().lower(): c for c in d['cps'] if c['type'] == 'service'}
max_cp_id   = max(c['id'] for c in d['cps'])

ws = wb['Услуги (подрядчики)']
excel_services = []
for r in range(3, ws.max_row + 1):
    name = ws.cell(r, 2).value
    if not name or not str(name).strip():
        continue
    excel_services.append({
        'name': str(name).strip(),
        'fio':  str(ws.cell(r, 3).value or '').strip() or '-',
        'ph':   str(ws.cell(r, 4).value or '').strip() or '-',
        'note': str(ws.cell(r, 5).value or '').strip() or '',
    })

added = updated = 0
for svc in excel_services:
    key = svc['name'].lower()
    if key in svc_by_name:
        cp = svc_by_name[key]
        cp['fio']  = svc['fio']
        cp['ph']   = svc['ph']
        cp['note'] = svc['note']
        updated += 1
    else:
        max_cp_id += 1
        new_cp = {
            'id': max_cp_id, 'n': svc['name'], 'org': svc['name'],
            'type': 'service', 'fio': svc['fio'], 'ph': svc['ph'], 'note': svc['note']
        }
        d['cps'].append(new_cp)
        svc_by_name[key] = new_cp
        added += 1

print(f"Услуги: обновлено {updated}, добавлено {added}")

# ── 2. СТАТЬИ РАСХОДОВ — читаем матрицу ролей ─────────────────────────────────
ROLES = ['director', 'accountant', 'supply', 'manager', 'constructor', 'production', 'brigadier']
# Колонки матрицы в листе: 5=Директор, 6=Бухгалтер, 7=Снабженец, 8=Менеджер, 9=Конструктор, 10=Произ-во, 11=Бригадир

ws8 = wb['Статьи расходов']
# Строим таблицу: name → список ролей (без директора, т.к. директор всегда видит всё)
role_map = {}   # item_name_lower → set of roles (excluding director)
for r in range(3, ws8.max_row + 1):
    name_val = ws8.cell(r, 1).value
    if not name_val or not str(name_val).strip():
        continue
    name = str(name_val).strip()
    if name.startswith('  ') or name.startswith('›') or name.startswith('✓'):
        continue  # skip group headers and legend
    roles = []
    for ci, role in enumerate(ROLES[1:], 6):  # skip director (col5), start from col6=accountant
        val = ws8.cell(r, ci).value
        if val and '✓' in str(val):
            roles.append(ROLES[ci - 5])  # offset: col6=ROLES[1]=accountant
    role_map[name.lower()] = roles

# Определяем корректное значение who
def compute_who(roles):
    """Возвращает строку или список для поля who."""
    all_non_dir = ['accountant', 'supply', 'manager', 'constructor', 'production', 'brigadier']
    if set(roles) >= set(all_non_dir):
        return 'all'
    if len(roles) == 0:
        return 'accountant'  # только директор — используем accountant как флаг
    if len(roles) == 1:
        return roles[0]
    return roles  # массив

item_changed = 0
for it in d.get('items', []):
    key = it['n'].strip().lower()
    if key not in role_map:
        continue
    new_who = compute_who(role_map[key])
    old_who = it.get('who')
    if new_who != old_who:
        it['who'] = new_who
        print(f"  [{it['id']}] {it['n']}: {old_who!r} → {new_who!r}")
        item_changed += 1

print(f"Статьи расходов: изменено {item_changed} записей")

# ── Сохранение ────────────────────────────────────────────────────────────────
tmp = DATA + '.tmp'
with open(tmp, 'w', encoding='utf-8') as f:
    json.dump(d, f, ensure_ascii=False, indent=2)
shutil.move(tmp, DATA)
print("\ndata.json сохранён.")
