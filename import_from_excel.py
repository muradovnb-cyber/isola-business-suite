"""
import_from_excel.py
Читает ISOLA_Справочник.xlsx и обновляет data.json.
Обновляются: Персонал, Поставщики, Клиенты, Услуги (подрядчики).
Транзакции, Заказы, Начисления — только для просмотра, не импортируются.
"""
import json, sys, os
sys.stdout.reconfigure(encoding='utf-8')

try:
    import openpyxl
except ImportError:
    print("Установите openpyxl: pip install openpyxl")
    sys.exit(1)

EXCEL_PATH = r"C:\Users\user\Desktop\ISOLA_Справочник_v2.xlsx"
DATA_PATH  = r"C:\Users\user\клод код\таблицы экзель\ISOLA_Business_Suite\isola-repo\data.json"

ROLE_EN = {
    'Директор': 'director', 'Бухгалтер': 'accountant',
    'Снабженец': 'supply', 'Снабжение': 'supply',
    'Менеджер': 'manager', 'Конструктор': 'constructor',
    'Производство': 'production', 'Бригадир': 'brigadier'
}

if not os.path.exists(EXCEL_PATH):
    print(f"Файл не найден: {EXCEL_PATH}")
    sys.exit(1)

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

def rows_from_sheet(ws, start_row=3):
    """Yield non-empty rows as lists, skip merged title/header rows."""
    headers = [ws.cell(2, c).value for c in range(1, ws.max_column+1)]
    for r in range(start_row, ws.max_row+1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column+1)]
        if any(v is not None for v in vals):
            yield dict(zip(headers, vals))

# Load current data.json
with open(DATA_PATH, 'r', encoding='utf-8-sig') as f:
    d = json.load(f)

# ── 1. ПЕРСОНАЛ ────────────────────────────────────────────────────────────────
if 'Персонал' in wb.sheetnames:
    ws = wb['Персонал']
    user_by_id = {u['id']: u for u in d['users']}
    max_id = max((u['id'] for u in d['users']), default=0)
    changed = 0
    for row in rows_from_sheet(ws):
        uid = row.get('ID')
        name = (row.get('ФИО') or '').strip()
        if not name:
            continue
        role_ru = (row.get('Роль') or '').strip()
        role_en = ROLE_EN.get(role_ru, role_ru.lower() if role_ru else 'supply')
        sal  = int(row.get('Оклад (сум)') or 0)
        comm = float(row.get('Комиссия %') or 0)
        ph   = str(row.get('Телефон') or '').strip()
        em   = str(row.get('Email') or '').strip()
        dept = str(row.get('Отдел') or '').strip()
        start= str(row.get('Дата найма') or '').strip()

        if uid and uid in user_by_id:
            u = user_by_id[uid]
            u['n']    = name
            u['role'] = role_en
            u['sal']  = sal
            u['comm'] = comm
            if ph:  u['ph']    = ph
            if em:  u['e']     = em
            if dept: u['dept'] = dept
            if start: u['start'] = start
        else:
            max_id += 1
            new_u = {'id': max_id, 'n': name, 'role': role_en,
                     'sal': sal, 'comm': comm, 'ph': ph, 'e': em,
                     'dept': dept, 'start': start, 'pwd': '1234'}
            d['users'].append(new_u)
            user_by_id[max_id] = new_u
        changed += 1
    print(f"Персонал: обновлено/добавлено {changed} записей")

# ── 2–4. КОНТРАГЕНТЫ ──────────────────────────────────────────────────────────
def import_cps(sheet_name, cp_type):
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    cp_by_id = {c['id']: c for c in d['cps']}
    max_id = max((c['id'] for c in d['cps']), default=0)
    changed = 0
    for row in rows_from_sheet(ws):
        cid  = row.get('ID')
        name = (row.get('Название') or '').strip()
        if not name:
            continue
        fio  = str(row.get('Контакт / ФИО') or '').strip()
        ph   = str(row.get('Телефон') or '').strip()
        note = str(row.get('Примечание / Категория') or row.get('Примечание') or '').strip()

        if cid and cid in cp_by_id:
            cp = cp_by_id[cid]
            cp['n']    = name
            cp['type'] = cp_type
            if fio: cp['fio'] = fio
            if ph:  cp['ph']  = ph
            if note: cp['note'] = note
        else:
            max_id += 1
            new_cp = {'id': max_id, 'n': name, 'type': cp_type,
                      'fio': fio, 'ph': ph, 'note': note}
            d['cps'].append(new_cp)
            cp_by_id[max_id] = new_cp
        changed += 1
    print(f"{sheet_name}: обновлено/добавлено {changed} записей")

import_cps('Поставщики',          'supplier')
import_cps('Клиенты',             'client')
import_cps('Услуги (подрядчики)', 'service')

# ── Сохраняем data.json ───────────────────────────────────────────────────────
# Пишем через временный файл для безопасности
import tempfile, shutil
tmp = DATA_PATH + '.tmp'
with open(tmp, 'w', encoding='utf-8') as f:
    json.dump(d, f, ensure_ascii=False, indent=2)
shutil.move(tmp, DATA_PATH)

print()
print("data.json обновлён:", DATA_PATH)
print("Откройте приложение (index.html) — данные загрузятся автоматически.")
