import json, sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DATA_PATH = r"C:\Users\user\клод код\таблицы экзель\ISOLA_Business_Suite\isola-repo\data.json"
OUT_PATH  = r"C:\Users\user\Desktop\ISOLA_Справочник_v2.xlsx"

with open(DATA_PATH, 'r', encoding='utf-8-sig') as f:
    d = json.load(f)

def hdr_fill(hex_bg):
    return PatternFill('solid', start_color=hex_bg)

def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def write_sheet(ws, title, headers, rows, bg_hex='1A6B3C', sub_hex='2E7D32'):
    ws.title = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    tc = ws.cell(1, 1, title)
    tc.font = Font(name='Arial', bold=True, size=13, color='FFFFFF')
    tc.fill = hdr_fill(bg_hex)
    tc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    for ci, h in enumerate(headers, 1):
        c = ws.cell(2, ci, h)
        c.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        c.fill = hdr_fill(sub_hex)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[2].height = 20

    for ri, row in enumerate(rows, 3):
        fill = hdr_fill('F9FBE7') if ri % 2 == 1 else hdr_fill('FFFFFF')
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val)
            c.font = Font(name='Arial', size=10)
            c.border = thin_border()
            c.fill = fill
            if isinstance(val, (int, float)) and val > 999:
                c.number_format = '#,##0'
                c.alignment = Alignment(horizontal='right', vertical='center')
            else:
                c.alignment = Alignment(vertical='center', wrap_text=True)
        ws.row_dimensions[ri].height = 18

user_map  = {u['id']: u['n'] for u in d['users']}
cp_map    = {c['id']: c['n'] for c in d['cps']}
item_map  = {it['id']: it['n'] for it in d.get('items', [])}
order_map = {o['id']: o['title'] for o in d.get('orders', [])}

ROLE_RU = {
    'director': 'Директор', 'accountant': 'Бухгалтер',
    'supply': 'Снабжение', 'manager': 'Менеджер',
    'constructor': 'Конструктор', 'production': 'Производство',
    'brigadier': 'Бригадир'
}
MN = ['Январь','Февраль','Март','Апрель','Май','Июнь',
      'Июль','Август','Сентябрь','Октябрь','Ноябрь','Декабрь']

wb = Workbook()

# === 1. ПЕРСОНАЛ ===
ws1 = wb.active
accrual_by_user = {}
for a in d.get('accruals', []):
    uid = a.get('eid') or a.get('empId')
    if uid:
        accrual_by_user[uid] = accrual_by_user.get(uid, 0) + (a.get('amt', 0) or 0)

rows_pers = []
for u in d['users']:
    rows_pers.append([
        u['id'], u['n'],
        ROLE_RU.get(u['role'], u['role']),
        u.get('dept') or '-',
        u.get('sal', 0) or 0,
        u.get('comm', 0) or 0,
        u.get('ph', '') or '',
        u.get('e', '') or '',
        u.get('start', '') or '',
        accrual_by_user.get(u['id'], 0),
    ])

write_sheet(ws1, 'Персонал',
    ['ID', 'ФИО', 'Роль', 'Отдел', 'Оклад (сум)', 'Комиссия %',
     'Телефон', 'Email', 'Дата найма', 'Начислено (сум)'],
    rows_pers, '1A6B3C', '2E7D32')
for col, w in zip('ABCDEFGHIJ', [5,26,16,16,16,12,20,26,14,18]):
    ws1.column_dimensions[col].width = w

# === 2. ПОСТАВЩИКИ ===
ws2 = wb.create_sheet()
suppliers = [c for c in d['cps'] if c['type'] == 'supplier']
rows_sup = [[c['id'], c['n'], c.get('fio','') or '-', c.get('ph','') or '-', c.get('note','') or '-']
            for c in suppliers]
write_sheet(ws2, 'Поставщики',
    ['ID', 'Название', 'Контакт / ФИО', 'Телефон', 'Примечание / Категория'],
    rows_sup, '1565C0', '1976D2')
for col, w in zip('ABCDE', [5,30,22,20,32]):
    ws2.column_dimensions[col].width = w

# === 3. УСЛУГИ (подрядчики) ===
ws3 = wb.create_sheet()
svc_cps = [c for c in d['cps'] if c['type'] == 'service']
if not svc_cps:
    svc_cps = [c for c in d['cps'] if 'услуг' in (c.get('note','') or '').lower()]
rows_svc = [[c['id'], c['n'], c.get('fio','') or '-', c.get('ph','') or '-', c.get('note','') or '-']
            for c in svc_cps]
write_sheet(ws3, 'Услуги (подрядчики)',
    ['ID', 'Название', 'Контакт / ФИО', 'Телефон', 'Примечание'],
    rows_svc, '6A1B9A', '7B1FA2')
for col, w in zip('ABCDE', [5,30,22,20,32]):
    ws3.column_dimensions[col].width = w

# === 4. КЛИЕНТЫ ===
ws4 = wb.create_sheet()
clients = [c for c in d['cps'] if c['type'] == 'client']
rows_cli = [[c['id'], c['n'], c.get('fio','') or '-', c.get('ph','') or '-', c.get('note','') or '-']
            for c in clients]
write_sheet(ws4, 'Клиенты',
    ['ID', 'Название', 'Контакт / ФИО', 'Телефон', 'Примечание'],
    rows_cli, '00695C', '00796B')
for col, w in zip('ABCDE', [5,30,22,20,32]):
    ws4.column_dimensions[col].width = w

# === 5. ЗАКАЗЫ ===
ws5 = wb.create_sheet()
STATUS_RU = {'active': 'Активен', 'cancelled': 'Отменён', 'done': 'Завершён', 'archive': 'Архив'}
rows_ord = []
for o in sorted(d['orders'], key=lambda x: x['id'], reverse=True):
    rows_ord.append([
        o['id'], o['title'],
        STATUS_RU.get(o.get('status', ''), o.get('status', '')),
        o.get('uzs', 0) or 0,
        o.get('usd', 0) or 0,
        user_map.get(o.get('mid'), '-'),
        o.get('created', '') or '',
        o.get('closed', '') or '-',
        o.get('note', '') or '-',
    ])
write_sheet(ws5, 'Заказы',
    ['ID', 'Название', 'Статус', 'Сумма (сум)', 'Сумма (USD)', 'Менеджер', 'Создан', 'Закрыт', 'Примечание'],
    rows_ord, 'E65100', 'F4511E')
for col, w in zip('ABCDEFGHI', [5,36,12,16,12,22,12,12,28]):
    ws5.column_dimensions[col].width = w

# === 6. ТРАНЗАКЦИИ ===
ws6 = wb.create_sheet()
TYPE_RU = {'income': 'Доход', 'expense': 'Расход', 'transfer': 'Перевод'}
rows_tx = []
for t in sorted(d['txs'], key=lambda x: x.get('date',''), reverse=True):
    rows_tx.append([
        t['id'],
        t.get('date', ''),
        TYPE_RU.get(t.get('type', ''), t.get('type', '')),
        item_map.get(t.get('iid'), '-'),
        cp_map.get(t.get('cpid'), '-'),
        order_map.get(t.get('oid'), '-'),
        t.get('uzs', 0) or 0,
        t.get('cur', 'UZS'),
        user_map.get(t.get('by'), '-'),
        t.get('note', '') or '-',
    ])
write_sheet(ws6, 'Транзакции',
    ['ID', 'Дата', 'Тип', 'Статья', 'Контрагент', 'Заказ', 'Сумма (сум)', 'Валюта', 'Кто', 'Примечание'],
    rows_tx, '37474F', '455A64')
for col, w in zip('ABCDEFGHIJ', [5,12,10,22,22,28,16,8,20,36]):
    ws6.column_dimensions[col].width = w

# === 7. НАЧИСЛЕНИЯ ===
ws7 = wb.create_sheet()
TYPE_ACC = {
    'salary': 'Оклад', 'commission': 'Комиссия',
    'bonus': 'Бонус', 'advance': 'Аванс', 'fine': 'Штраф', 'payout': 'Выплата'
}
rows_acc = []
for a in d['accruals']:
    uid = a.get('eid') or a.get('empId')
    mo = a.get('mo') or a.get('month')
    yr = a.get('yr') or a.get('year')
    mo_str = MN[mo-1] if mo and 1 <= mo <= 12 else '-'
    rows_acc.append([
        a['id'],
        a.get('date', ''),
        user_map.get(uid, '-'),
        TYPE_ACC.get(a.get('type', ''), a.get('type', '')),
        a.get('amt', 0) or 0,
        mo_str,
        yr or '-',
        a.get('note', '') or '-',
    ])
write_sheet(ws7, 'Начисления ЗП',
    ['ID', 'Дата', 'Сотрудник', 'Тип', 'Сумма (сум)', 'Месяц', 'Год', 'Примечание'],
    rows_acc, '1A6B3C', '2E7D32')
for col, w in zip('ABCDEFGH', [5,12,26,14,16,14,8,40]):
    ws7.column_dimensions[col].width = w

# === 8. СТАТЬИ РАСХОДОВ (матрица ролей) ===
ws8 = wb.create_sheet()
ws8.title = 'Статьи расходов'

CAT_RU = {'gen': 'Общие', 'ord': 'По заказу'}
SUB_RU = {
    'salary': 'Зарплаты', 'office': 'Офис/Аренда', 'bank': 'Банк/Обнал',
    'other': 'Прочее', 'mat': 'Материалы', 'trans': 'Транспорт',
    'food': 'Питание', 'work': 'Работа бригады', 'comm': 'Комиссия',
    'svc': 'Услуги подрядч.', 'warehouse': 'Склад'
}

# Role visibility: who field determines who sees the item
# 'all' = everyone, else specific role
# Director always sees everything
ROLES = ['director', 'accountant', 'supply', 'manager', 'constructor', 'production', 'brigadier']
ROLE_LABELS = ['Директор', 'Бухгалтер', 'Снабженец', 'Менеджер', 'Конструктор', 'Произ-во', 'Бригадир']

def role_visible(item, role):
    who = item.get('who', 'all')
    if role == 'director':
        return True
    return who == 'all' or who == role

# Title row
ws8.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4+len(ROLES))
tc = ws8.cell(1, 1, 'Статьи расходов — Видимость по ролям')
tc.font = Font(name='Arial', bold=True, size=13, color='FFFFFF')
tc.fill = hdr_fill('37474F')
tc.alignment = Alignment(horizontal='center', vertical='center')
ws8.row_dimensions[1].height = 26

# Header row
hdrs = ['Название', 'Категория', 'Подкатегория', 'Тип записи'] + ROLE_LABELS
for ci, h in enumerate(hdrs, 1):
    c = ws8.cell(2, ci, h)
    c.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    c.fill = hdr_fill('455A64')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = thin_border()
ws8.row_dimensions[2].height = 22

# Separate items by category for better readability
gen_items = [it for it in d.get('items', []) if it.get('cat') == 'gen']
ord_items = [it for it in d.get('items', []) if it.get('cat') == 'ord']
all_items = gen_items + ord_items

# Color fills for checkmark cells
YES_FILL = hdr_fill('C8E6C9')   # light green = visible
NO_FILL  = hdr_fill('FFEBEE')   # light red = not visible
CAT_GEN_FILL = hdr_fill('E3F2FD')   # light blue for gen rows
CAT_ORD_FILL = hdr_fill('FFF8E1')   # light amber for ord rows

# Group header rows
def add_group_header(ws, row, label, bg):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4+len(ROLES))
    c = ws.cell(row, 1, label)
    c.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    c.fill = hdr_fill(bg)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 20

ri = 3
add_group_header(ws8, ri, '  ОБЩИЕ РАСХОДЫ (видны в разделе "Расходы" / "Транзакции")', '1565C0')
ri += 1

for it in gen_items:
    base_fill = CAT_GEN_FILL
    row_vals = [
        it['n'],
        CAT_RU.get(it.get('cat',''), it.get('cat','')),
        SUB_RU.get(it.get('sub',''), it.get('sub','')),
        'Общая транзакция'
    ]
    for ci, val in enumerate(row_vals, 1):
        c = ws8.cell(ri, ci, val)
        c.font = Font(name='Arial', size=10)
        c.border = thin_border()
        c.fill = base_fill
        c.alignment = Alignment(vertical='center', wrap_text=True)
    # Role columns
    for rj, role in enumerate(ROLES):
        vis = role_visible(it, role)
        col = 5 + rj
        c = ws8.cell(ri, col, '✓' if vis else '—')
        c.font = Font(name='Arial', size=11, bold=vis,
                      color='1B5E20' if vis else '999999')
        c.fill = YES_FILL if vis else NO_FILL
        c.border = thin_border()
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws8.row_dimensions[ri].height = 18
    ri += 1

add_group_header(ws8, ri, '  РАСХОДЫ ПО ЗАКАЗУ (видны в карточке заказа)', 'E65100')
ri += 1

# Sub-groups within order items
svc_subs = {}
for it in ord_items:
    sub = it.get('sub', 'other')
    svc_subs.setdefault(sub, []).append(it)

SUB_ORDER = ['mat', 'trans', 'food', 'work', 'comm', 'svc', 'other']
for sub in SUB_ORDER:
    items_in_sub = svc_subs.get(sub, [])
    if not items_in_sub:
        continue
    # Sub-group mini-header
    ws8.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=4+len(ROLES))
    c = ws8.cell(ri, 1, f'    › {SUB_RU.get(sub, sub)}')
    c.font = Font(name='Arial', bold=True, size=10, italic=True, color='424242')
    c.fill = hdr_fill('ECEFF1')
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws8.row_dimensions[ri].height = 16
    ri += 1

    for it in items_in_sub:
        row_vals = [
            it['n'],
            CAT_RU.get(it.get('cat',''), it.get('cat','')),
            SUB_RU.get(it.get('sub',''), it.get('sub','')),
            'Расход по заказу'
        ]
        for ci, val in enumerate(row_vals, 1):
            c = ws8.cell(ri, ci, val)
            c.font = Font(name='Arial', size=10)
            c.border = thin_border()
            c.fill = CAT_ORD_FILL
            c.alignment = Alignment(vertical='center', wrap_text=True)
        for rj, role in enumerate(ROLES):
            vis = role_visible(it, role)
            col = 5 + rj
            c = ws8.cell(ri, col, '✓' if vis else '—')
            c.font = Font(name='Arial', size=11, bold=vis,
                          color='1B5E20' if vis else '999999')
            c.fill = YES_FILL if vis else NO_FILL
            c.border = thin_border()
            c.alignment = Alignment(horizontal='center', vertical='center')
        ws8.row_dimensions[ri].height = 18
        ri += 1

# Legend row
ri += 1
ws8.cell(ri, 1, '✓ = роль видит эту статью расходов').font = Font(name='Arial', size=9, color='1B5E20')
ws8.cell(ri, 2, '— = статья скрыта от этой роли').font = Font(name='Arial', size=9, color='999999')

# Column widths
col_widths = [30, 14, 18, 18] + [12]*len(ROLES)
for ci, w in enumerate(col_widths, 1):
    from openpyxl.utils import get_column_letter
    ws8.column_dimensions[get_column_letter(ci)].width = w

# === 9. ИНСТРУКЦИЯ ПО ИМПОРТУ ===
ws9 = wb.create_sheet()
ws9.title = 'Инструкция'
ws9.merge_cells('A1:E1')
c = ws9.cell(1, 1, 'Инструкция: как загрузить данные из Excel в приложение')
c.font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
c.fill = hdr_fill('1A6B3C')
c.alignment = Alignment(horizontal='center', vertical='center')
ws9.row_dimensions[1].height = 30

instructions = [
    ('1. Заполните нужные листы', 'Персонал, Поставщики, Клиенты, Услуги — можно редактировать название, телефон, оклад и т.д.'),
    ('2. Не меняйте столбец ID', 'ID используется для сопоставления записей. Если добавляете новую строку — оставьте ID пустым или укажите новый уникальный номер.'),
    ('3. Сохраните файл', 'Сохраните Excel как ISOLA_Справочник.xlsx в той же папке или на рабочем столе.'),
    ('4. Запустите скрипт импорта', 'Откройте PowerShell и выполните:\n  python "C:\\Users\\user\\клод код\\таблицы экзель\\ISOLA_Business_Suite\\isola-repo\\import_from_excel.py"'),
    ('5. Скрипт обновит data.json', 'Данные из листов Персонал / Поставщики / Клиенты / Услуги будут записаны в data.json. Приложение автоматически подхватит изменения при следующем открытии.'),
    ('6. Что НЕ импортируется', 'Транзакции, Начисления, Заказы — только для просмотра. Их изменяйте через интерфейс приложения, не через Excel.'),
    ('', ''),
    ('Листы для редактирования:', ''),
    ('  • Персонал', 'ФИО, роль, оклад, телефон, email, дата найма'),
    ('  • Поставщики', 'Название, контакт, телефон, примечание'),
    ('  • Клиенты', 'Название, контакт, телефон, примечание'),
    ('  • Услуги (подрядчики)', 'Название, контакт, телефон, примечание'),
]
for ri2, (step, detail) in enumerate(instructions, 3):
    c1 = ws9.cell(ri2, 1, step)
    c1.font = Font(name='Arial', bold=True, size=10)
    c1.alignment = Alignment(vertical='top', wrap_text=True)
    c2 = ws9.cell(ri2, 2, detail)
    c2.font = Font(name='Arial', size=10)
    c2.alignment = Alignment(vertical='top', wrap_text=True)
    ws9.row_dimensions[ri2].height = 30
    ws9.merge_cells(start_row=ri2, start_column=2, end_row=ri2, end_column=5)

ws9.column_dimensions['A'].width = 32
ws9.column_dimensions['B'].width = 70

wb.save(OUT_PATH)
print('OK:', OUT_PATH)
